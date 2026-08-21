"""
Router de Reportes SENNOVA
Generación de reportes consolidados (Excel / CSV) para seguimiento institucional y reportes trimestrales nacionales.
"""

from io import BytesIO, StringIO
from datetime import datetime, timezone
import uuid
from typing import Optional, Literal

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy.exc import OperationalError, SQLAlchemyError
from sqlalchemy import func

from app.database import get_db
from app.auth import get_current_admin, get_current_investigador_or_instructor
from app.models import User, Proyecto, Grupo, Semillero, Producto

# Importar librerías de Excel
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

router = APIRouter(
    prefix="/reportes",
    tags=["Reportes SENNOVA"]
)

# Paleta institucional de estilos Excel (SENA / SENNOVA)
HEADER_FILL = PatternFill(start_color="047857", end_color="047857", fill_type="solid")  # Emerald 700
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(bold=True, size=13, color="0F172A")
SUBTITLE_FONT = Font(size=9, color="64748B")
TOTAL_FILL = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
TOTAL_FONT = Font(bold=True, size=9, color="0F172A")

THIN_BORDER = Border(
    left=Side(style='thin', color='CBD5E1'),
    right=Side(style='thin', color='CBD5E1'),
    top=Side(style='thin', color='CBD5E1'),
    bottom=Side(style='thin', color='CBD5E1')
)

TOTAL_BORDER = Border(
    left=Side(style='thin', color='CBD5E1'),
    right=Side(style='thin', color='CBD5E1'),
    top=Side(style='medium', color='047857'),
    bottom=Side(style='double', color='047857')
)


def get_estado_color(estado: str) -> str:
    """Retorna color de celda según el estado del proyecto."""
    colors = {
        "Aprobado": "D1FAE5",      # Emerald 100
        "En ejecución": "DBEAFE",  # Blue 100
        "Finalizado": "E2E8F0",    # Slate 200
        "Rechazado": "FEE2E2",     # Red 100
        "Formulación": "FEF3C7",   # Amber 100
        "Enviado": "E0E7FF"        # Indigo 100
    }
    return colors.get(str(estado or '').strip(), "FFFFFF")


def auto_fit_columns(ws, max_cols, min_width=12, max_width=45):
    """Ajusta automáticamente el ancho de las columnas con límites mínimos y máximos."""
    for col_idx in range(1, max_cols + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[col_letter]:
            if cell.row > 4 and cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 3, max_width))


# ─── 1. Consolidado de Proyectos ─────────────────────────────────────────────

@router.get("/proyectos-consolidado")
def generar_consolidado_proyectos(
    año: Optional[int] = Query(None, description="Año de vigencia de los proyectos"),
    formato: Literal["excel", "csv"] = Query("excel", description="Formato de salida"),
    current_user: User = Depends(get_current_investigador_or_instructor),
    db: Session = Depends(get_db)
):
    """
    Genera reporte consolidado de proyectos para seguimiento y reportes nacionales SENNOVA.
    Incluye: proyectos, equipo, presupuesto asignado, productos asociados y líneas.
    """
    if not EXCEL_AVAILABLE and formato == "excel":
        raise HTTPException(
            status_code=500,
            detail="openpyxl no está instalado. Ejecuta: pip install openpyxl"
        )
    
    try:
        query = db.query(Proyecto)
        if año:
            query = query.filter(Proyecto.vigencia == año)
        proyectos = query.all()
        
        if formato == "excel":
            return _generar_excel_consolidado(proyectos, db, año)
        else:
            return _generar_csv_consolidado(proyectos, año)
    except (OperationalError, SQLAlchemyError) as db_err:
        raise db_err
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


def _generar_excel_consolidado(proyectos, db, año_filtro):
    """Genera archivo Excel de proyectos con formato institucional SENNOVA."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Consolidado Proyectos"
    
    # Encabezado institucional
    año_texto = str(año_filtro) if año_filtro else "Todas las vigencias"
    ws["A1"] = "CONSOLIDADO DE PROYECTOS DE INVESTIGACIÓN - SENNOVA CGAO VÉLEZ"
    ws["A2"] = f"Vigencia: {año_texto} | Total Proyectos: {len(proyectos)} | Generado: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"
    
    ws.merge_cells("A1:K1")
    ws.merge_cells("A2:K2")
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A2"].font = SUBTITLE_FONT
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    
    headers = [
        "Código SGPS", "Nombre Corto", "Nombre Completo del Proyecto", "Estado", 
        "Vigencia", "Presupuesto Total (COP)", "Convocatoria", "Investigador Líder",
        "N° Miembros", "N° Productos", "Tipología / Línea"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    
    row = 5
    total_presupuesto = 0
    total_miembros = 0
    total_productos = 0
    
    for proyecto in proyectos:
        n_miembros = len(proyecto.equipo) if proyecto.equipo else 0
        n_productos = len(proyecto.productos) if hasattr(proyecto, 'productos') and proyecto.productos else 0
        lider = proyecto.owner.nombre if (proyecto.owner and proyecto.owner.nombre) else "Sin asignar"
        convocatoria = proyecto.convocatoria.nombre if (proyecto.convocatoria and proyecto.convocatoria.nombre) else "N/A"
        presupuesto = float(proyecto.presupuesto_total or 0)
        
        total_presupuesto += presupuesto
        total_miembros += n_miembros
        total_productos += n_productos
        
        nombre_proyecto = proyecto.nombre or "Sin nombre"
        nombre_corto = proyecto.nombre_corto or (nombre_proyecto[:45] if nombre_proyecto else "Sin definir")
        
        data = [
            proyecto.codigo_sgps or "Sin código",
            nombre_corto,
            nombre_proyecto,
            proyecto.estado or "Aprobado",
            proyecto.vigencia or "N/A",
            presupuesto,
            convocatoria,
            lider,
            n_miembros,
            n_productos,
            proyecto.tipologia or proyecto.linea_programatica or "I+D"
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = THIN_BORDER
            cell.font = Font(size=9)
            
            if col in [2, 3, 7, 8, 11]:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif col == 6:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = '"$"#,##0'
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            if col == 4:
                color_hex = get_estado_color(str(value))
                cell.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
        
        row += 1
    
    # Fila de Totales
    total_row = row
    ws.cell(row=total_row, column=1, value="TOTALES").font = TOTAL_FONT
    ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=total_row, column=1).fill = TOTAL_FILL
    ws.cell(row=total_row, column=1).border = TOTAL_BORDER
    
    for c in range(2, 12):
        cell = ws.cell(row=total_row, column=c)
        cell.fill = TOTAL_FILL
        cell.border = TOTAL_BORDER
        cell.font = TOTAL_FONT
        if c == 6:
            cell.value = total_presupuesto
            cell.number_format = '"$"#,##0'
            cell.alignment = Alignment(horizontal="right", vertical="center")
        elif c == 9:
            cell.value = total_miembros
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif c == 10:
            cell.value = total_productos
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.value = ""
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    auto_fit_columns(ws, len(headers))
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"consolidado_proyectos_{año_filtro or 'todos'}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


def _generar_csv_consolidado(proyectos, año_filtro):
    """Genera archivo CSV compatible con Excel."""
    import csv
    str_output = StringIO()
    
    headers = [
        "codigo_sgps", "nombre_corto", "nombre_completo", "estado", 
        "vigencia", "presupuesto_total_cop", "convocatoria", "lider",
        "num_miembros", "num_productos", "tipologia"
    ]
    
    writer = csv.writer(str_output)
    writer.writerow(headers)
    
    for proyecto in proyectos:
        n_miembros = len(proyecto.equipo) if proyecto.equipo else 0
        n_productos = len(proyecto.productos) if hasattr(proyecto, 'productos') and proyecto.productos else 0
        lider = proyecto.owner.nombre if (proyecto.owner and proyecto.owner.nombre) else "Sin asignar"
        convocatoria = proyecto.convocatoria.nombre if (proyecto.convocatoria and proyecto.convocatoria.nombre) else "N/A"
        nombre_proyecto = proyecto.nombre or "Sin nombre"
        nombre_corto = proyecto.nombre_corto or (nombre_proyecto[:45] if nombre_proyecto else "")
        
        writer.writerow([
            proyecto.codigo_sgps or "",
            nombre_corto,
            nombre_proyecto,
            proyecto.estado or "Aprobado",
            proyecto.vigencia or "",
            proyecto.presupuesto_total or 0,
            convocatoria,
            lider,
            n_miembros,
            n_productos,
            proyecto.tipologia or proyecto.linea_programatica or ""
        ])
    
    filename = f"consolidado_proyectos_{año_filtro or 'todos'}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    output = BytesIO(str_output.getvalue().encode('utf-8-sig'))
    
    return StreamingResponse(
        output,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ─── 2. Consolidado de Grupos de Investigación ───────────────────────────────

@router.get("/grupos-consolidado")
def generar_consolidado_grupos(
    formato: Literal["excel", "csv"] = Query("excel"),
    current_user: User = Depends(get_current_investigador_or_instructor),
    db: Session = Depends(get_db)
):
    """Genera reporte consolidado de grupos de investigación (GrupLAC)."""
    if not EXCEL_AVAILABLE and formato == "excel":
        raise HTTPException(status_code=500, detail="openpyxl no está instalado")
    
    try:
        grupos = db.query(Grupo).all()
        
        if formato == "excel":
            wb = Workbook()
            ws = wb.active
            ws.title = "Grupos GRUPLAC"
            
            # Título
            ws["A1"] = "GRUPOS DE INVESTIGACIÓN (GRUPLAC) - SENNOVA CGAO VÉLEZ"
            ws["A2"] = f"Total Grupos: {len(grupos)} | Generado: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"
            ws.merge_cells("A1:H1")
            ws.merge_cells("A2:H2")
            ws["A1"].font = TITLE_FONT
            ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
            ws["A2"].font = SUBTITLE_FONT
            ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
            
            headers = [
                "Nombre del Grupo", "Código GrupLAC", "Clasificación MinCiencias", "Líder / Director", 
                "N° Integrantes", "N° Semilleros", "Fecha Creación", "Líneas de Investigación"
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col, value=header)
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = THIN_BORDER
            
            row = 5
            total_integrantes = 0
            total_semilleros = 0
            
            for grupo in grupos:
                n_integrantes = len(grupo.integrantes) if grupo.integrantes else 0
                n_semilleros = len(grupo.semilleros) if hasattr(grupo, 'semilleros') and grupo.semilleros else 0
                lineas = ", ".join(grupo.lineas_investigacion or [])
                fecha_creacion = grupo.created_at.strftime('%Y-%m-%d') if grupo.created_at else "N/A"
                
                total_integrantes += n_integrantes
                total_semilleros += n_semilleros
                
                data = [
                    grupo.nombre,
                    grupo.codigo_gruplac or "N/A",
                    grupo.clasificacion or "Reconocido",
                    grupo.owner.nombre if (grupo.owner and grupo.owner.nombre) else "Sin líder",
                    n_integrantes,
                    n_semilleros,
                    fecha_creacion,
                    lineas
                ]
                
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = THIN_BORDER
                    cell.font = Font(size=9)
                    cell.alignment = Alignment(horizontal="left" if col in [1, 4, 8] else "center", vertical="center")
                
                row += 1
            
            # Totales
            total_row = row
            ws.cell(row=total_row, column=1, value="TOTALES").font = TOTAL_FONT
            ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=total_row, column=1).fill = TOTAL_FILL
            ws.cell(row=total_row, column=1).border = TOTAL_BORDER
            
            for c in range(2, 9):
                cell = ws.cell(row=total_row, column=c)
                cell.fill = TOTAL_FILL
                cell.border = TOTAL_BORDER
                cell.font = TOTAL_FONT
                if c == 5:
                    cell.value = total_integrantes
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif c == 6:
                    cell.value = total_semilleros
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.value = ""
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            
            auto_fit_columns(ws, len(headers))
            
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename=consolidado_grupos_{datetime.now().strftime('%Y%m%d')}.xlsx"}
            )
        else:
            import csv
            str_output = StringIO()
            writer = csv.writer(str_output)
            writer.writerow(["nombre", "codigo_gruplac", "clasificacion", "lider", "integrantes", "semilleros", "fecha_creacion", "lineas"])
            
            for grupo in grupos:
                writer.writerow([
                    grupo.nombre,
                    grupo.codigo_gruplac or "",
                    grupo.clasificacion or "",
                    grupo.owner.nombre if (grupo.owner and grupo.owner.nombre) else "",
                    len(grupo.integrantes) if grupo.integrantes else 0,
                    len(grupo.semilleros) if hasattr(grupo, 'semilleros') and grupo.semilleros else 0,
                    grupo.created_at.strftime('%Y-%m-%d') if grupo.created_at else "",
                    ", ".join(grupo.lineas_investigacion or [])
                ])
            
            output = BytesIO(str_output.getvalue().encode('utf-8-sig'))
            return StreamingResponse(
                output,
                media_type="text/csv; charset=utf-8",
                headers={"Content-Disposition": f"attachment; filename=consolidado_grupos_{datetime.now().strftime('%Y%m%d')}.csv"}
            )
    except (OperationalError, SQLAlchemyError) as db_err:
        raise db_err
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ─── 3. Consolidado de Productos de Investigación ────────────────────────────

@router.get("/productos-consolidado")
def generar_consolidado_productos(
    año: Optional[int] = Query(None, description="Año de publicación"),
    verificados_only: bool = Query(False, description="Solo productos verificados"),
    formato: Literal["excel", "csv"] = Query("excel"),
    current_user: User = Depends(get_current_investigador_or_instructor),
    db: Session = Depends(get_db)
):
    """Genera reporte consolidado de productos CTeI para convocatorias MinCiencias."""
    if not EXCEL_AVAILABLE and formato == "excel":
        raise HTTPException(status_code=500, detail="openpyxl no está instalado")
    
    try:
        query = db.query(Producto)
        if año:
            query = query.filter(
                Producto.fecha_publicacion >= f"{año}-01-01", 
                Producto.fecha_publicacion <= f"{año}-12-31"
            )
        if verificados_only:
            query = query.filter(Producto.is_verificado == True)
        
        productos = query.all()
        
        if formato == "excel":
            wb = Workbook()
            ws = wb.active
            ws.title = "Productos MinCiencias"
            
            ws["A1"] = "PRODUCTOS DE INVESTIGACIÓN Y DESARROLLO (CTeI) - SENNOVA CGAO VÉLEZ"
            filtro_texto = f"Año: {año}" if año else "Todos los años"
            filtro_texto += " | Solo verificados" if verificados_only else " | Todos los estados"
            ws["A2"] = f"{filtro_texto} | Total Productos: {len(productos)} | Generado: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"
            
            ws.merge_cells("A1:I1")
            ws.merge_cells("A2:I2")
            ws["A1"].font = TITLE_FONT
            ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
            ws["A2"].font = SUBTITLE_FONT
            ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
            
            headers = [
                "Tipo MinCiencias", "Nombre del Producto", "Descripción", "Fecha Publicación", 
                "DOI / Identificador", "Verificado", "Proyecto Asociado", "Autor / Investigador", "URL / Soporte"
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col, value=header)
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = THIN_BORDER
            
            row = 5
            total_verificados = 0
            
            for producto in productos:
                if producto.is_verificado:
                    total_verificados += 1
                    
                proyecto_nombre = producto.proyecto.nombre_corto if (producto.proyecto and producto.proyecto.nombre_corto) else (producto.proyecto.nombre if producto.proyecto else "Sin proyecto")
                autor_nombre = producto.owner.nombre if (producto.owner and producto.owner.nombre) else "N/A"
                fecha_pub = str(producto.fecha_publicacion or "N/A")
                
                data = [
                    producto.tipo,
                    producto.nombre,
                    producto.descripcion or "",
                    fecha_pub,
                    producto.doi or "N/A",
                    "SÍ" if producto.is_verificado else "NO",
                    proyecto_nombre,
                    autor_nombre,
                    producto.url or "N/A"
                ]
                
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = THIN_BORDER
                    cell.font = Font(size=9)
                    cell.alignment = Alignment(horizontal="left" if col in [2, 3, 7, 8, 9] else "center", vertical="center")
                    
                    if col == 6:
                        color_hex = "D1FAE5" if producto.is_verificado else "FEE2E2"
                        cell.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
                
                row += 1
            
            # Fila de Resumen
            total_row = row
            ws.cell(row=total_row, column=1, value="RESUMEN").font = TOTAL_FONT
            ws.cell(row=total_row, column=1).fill = TOTAL_FILL
            ws.cell(row=total_row, column=1).border = TOTAL_BORDER
            ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
            
            for c in range(2, 10):
                cell = ws.cell(row=total_row, column=c)
                cell.fill = TOTAL_FILL
                cell.border = TOTAL_BORDER
                cell.font = TOTAL_FONT
                if c == 2:
                    cell.value = f"Total: {len(productos)} productos"
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                elif c == 6:
                    cell.value = f"{total_verificados} Verificados"
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.value = ""
            
            auto_fit_columns(ws, len(headers))
            
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            suffix = f"{año or 'todos'}_{'verificados' if verificados_only else 'todos'}"
            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename=consolidado_productos_{suffix}_{datetime.now().strftime('%Y%m%d')}.xlsx"}
            )
        else:
            import csv
            str_output = StringIO()
            writer = csv.writer(str_output)
            writer.writerow(["tipo", "nombre", "descripcion", "fecha_publicacion", "doi", "verificado", "proyecto", "autor", "url"])
            
            for producto in productos:
                proyecto_nombre = producto.proyecto.nombre_corto if (producto.proyecto and producto.proyecto.nombre_corto) else (producto.proyecto.nombre if producto.proyecto else "")
                writer.writerow([
                    producto.tipo,
                    producto.nombre,
                    producto.descripcion or "",
                    str(producto.fecha_publicacion or ""),
                    producto.doi or "",
                    "Sí" if producto.is_verificado else "No",
                    proyecto_nombre,
                    producto.owner.nombre if (producto.owner and producto.owner.nombre) else "",
                    producto.url or ""
                ])
            
            suffix = f"{año or 'todos'}_{'verificados' if verificados_only else 'todos'}"
            output = BytesIO(str_output.getvalue().encode('utf-8-sig'))
            return StreamingResponse(
                output,
                media_type="text/csv; charset=utf-8",
                headers={"Content-Disposition": f"attachment; filename=consolidado_productos_{suffix}_{datetime.now().strftime('%Y%m%d')}.csv"}
            )
    except (OperationalError, SQLAlchemyError) as db_err:
        raise db_err
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ─── 4. Consolidado de Semilleros de Investigación ───────────────────────────

@router.get("/semilleros-consolidado")
def generar_consolidado_semilleros(
    formato: Literal["excel", "csv"] = Query("excel"),
    current_user: User = Depends(get_current_investigador_or_instructor),
    db: Session = Depends(get_db)
):
    """Genera reporte consolidado de semilleros y aprendices vinculados."""
    if not EXCEL_AVAILABLE and formato == "excel":
        raise HTTPException(status_code=500, detail="openpyxl no está instalado")
    
    try:
        semilleros = db.query(Semillero).all()
        
        if formato == "excel":
            wb = Workbook()
            ws = wb.active
            ws.title = "Semilleros y Aprendices"
            
            ws["A1"] = "SEMILLEROS DE INVESTIGACIÓN FORMATIVA - SENNOVA CGAO VÉLEZ"
            ws["A2"] = f"Total Semilleros: {len(semilleros)} | Generado: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"
            ws.merge_cells("A1:G1")
            ws.merge_cells("A2:G2")
            ws["A1"].font = TITLE_FONT
            ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
            ws["A2"].font = SUBTITLE_FONT
            ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
            
            headers = [
                "Nombre del Semillero", "Sigla / Código", "Grupo Vinculado", "Líder de Semillero", 
                "Fecha Creación", "N° Aprendices", "Estado"
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col, value=header)
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = THIN_BORDER
            
            row = 5
            total_aprendices = 0
            
            for semillero in semilleros:
                n_aprendices = len(semillero.aprendices) if hasattr(semillero, 'aprendices') and semillero.aprendices else 0
                grupo_nombre = semillero.grupo.nombre if (semillero.grupo and semillero.grupo.nombre) else "Sin grupo"
                lider_nombre = semillero.owner.nombre if (semillero.owner and semillero.owner.nombre) else "Sin líder"
                fecha_creacion = semillero.created_at.strftime('%Y-%m-%d') if semillero.created_at else "N/A"
                
                total_aprendices += n_aprendices
                
                data = [
                    semillero.nombre,
                    getattr(semillero, 'sigla', None) or getattr(semillero, 'codigo', None) or "N/A",
                    grupo_nombre,
                    lider_nombre,
                    fecha_creacion,
                    n_aprendices,
                    semillero.estado or "Activo"
                ]
                
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = THIN_BORDER
                    cell.font = Font(size=9)
                    cell.alignment = Alignment(horizontal="left" if col in [1, 3, 4] else "center", vertical="center")
                
                row += 1
            
            # Totales
            total_row = row
            ws.cell(row=total_row, column=1, value="TOTALES").font = TOTAL_FONT
            ws.cell(row=total_row, column=1).fill = TOTAL_FILL
            ws.cell(row=total_row, column=1).border = TOTAL_BORDER
            ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
            
            for c in range(2, 8):
                cell = ws.cell(row=total_row, column=c)
                cell.fill = TOTAL_FILL
                cell.border = TOTAL_BORDER
                cell.font = TOTAL_FONT
                if c == 6:
                    cell.value = total_aprendices
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.value = ""
            
            auto_fit_columns(ws, len(headers))
            
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename=consolidado_semilleros_{datetime.now().strftime('%Y%m%d')}.xlsx"}
            )
        else:
            import csv
            str_output = StringIO()
            writer = csv.writer(str_output)
            writer.writerow(["nombre", "sigla", "grupo", "lider", "fecha_creacion", "aprendices", "estado"])
            
            for semillero in semilleros:
                writer.writerow([
                    semillero.nombre,
                    getattr(semillero, 'sigla', None) or getattr(semillero, 'codigo', None) or "",
                    semillero.grupo.nombre if (semillero.grupo and semillero.grupo.nombre) else "",
                    semillero.owner.nombre if (semillero.owner and semillero.owner.nombre) else "",
                    semillero.created_at.strftime('%Y-%m-%d') if semillero.created_at else "",
                    len(semillero.aprendices) if hasattr(semillero, 'aprendices') and semillero.aprendices else 0,
                    semillero.estado or "Activo"
                ])
            
            output = BytesIO(str_output.getvalue().encode('utf-8-sig'))
            return StreamingResponse(
                output,
                media_type="text/csv; charset=utf-8",
                headers={"Content-Disposition": f"attachment; filename=consolidado_semilleros_{datetime.now().strftime('%Y%m%d')}.csv"}
            )
    except (OperationalError, SQLAlchemyError) as db_err:
        raise db_err
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ─── 5. Consolidado de Talento Humano e Investigadores ───────────────────────

@router.get("/talento-consolidado")
def generar_consolidado_talento(
    formato: Literal["excel", "csv"] = Query("excel"),
    current_user: User = Depends(get_current_investigador_or_instructor),
    db: Session = Depends(get_db)
):
    """Genera reporte consolidado de talento humano, investigadores e instructores."""
    if not EXCEL_AVAILABLE and formato == "excel":
        raise HTTPException(status_code=500, detail="openpyxl no está instalado")

    try:
        investigadores = db.query(User).filter(User.rol.in_(['investigador', 'instructor', 'admin'])).all()
        
        if formato == "excel":
            wb = Workbook()
            ws = wb.active
            ws.title = "Talento SENNOVA"
            
            ws["A1"] = "DIRECTORIO CONSOLIDADO DE TALENTO HUMANO - SENNOVA CGAO"
            ws["A2"] = f"Total Personal CTeI: {len(investigadores)} | Generado: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"
            ws.merge_cells("A1:G1")
            ws.merge_cells("A2:G2")
            ws["A1"].font = TITLE_FONT
            ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
            ws["A2"].font = SUBTITLE_FONT
            ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
            
            headers = ["Nombre Completo", "Correo Institucional", "Regional", "Sede / Centro", "Nivel Académico", "Rol SENNOVA", "Estado"]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col, value=header)
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = THIN_BORDER
                
            row = 5
            total_activos = 0
            for inv in investigadores:
                if inv.is_active:
                    total_activos += 1
                
                data = [
                    inv.nombre,
                    inv.email,
                    inv.regional or "SANTANDER",
                    inv.sede or "CGAO VÉLEZ",
                    inv.nivel_academico or "Profesional",
                    (inv.rol or 'investigador').capitalize(),
                    "Activo" if inv.is_active else "Inactivo"
                ]
                
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = THIN_BORDER
                    cell.font = Font(size=9)
                    cell.alignment = Alignment(horizontal="left" if col in [1, 2, 5] else "center", vertical="center")
                    
                    if col == 7:
                        color_hex = "D1FAE5" if inv.is_active else "FEE2E2"
                        cell.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
                
                row += 1
            
            # Totales
            total_row = row
            ws.cell(row=total_row, column=1, value="TOTALES").font = TOTAL_FONT
            ws.cell(row=total_row, column=1).fill = TOTAL_FILL
            ws.cell(row=total_row, column=1).border = TOTAL_BORDER
            ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
            
            for c in range(2, 8):
                cell = ws.cell(row=total_row, column=c)
                cell.fill = TOTAL_FILL
                cell.border = TOTAL_BORDER
                cell.font = TOTAL_FONT
                if c == 2:
                    cell.value = f"Total: {len(investigadores)} registrados"
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                elif c == 7:
                    cell.value = f"{total_activos} Activos"
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.value = ""
            
            auto_fit_columns(ws, len(headers))
            
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": "attachment; filename=reporte_talento_sennova.xlsx"}
            )
        else:
            import csv
            str_output = StringIO()
            writer = csv.writer(str_output)
            writer.writerow(["nombre", "email", "regional", "sede", "nivel_academico", "rol", "estado"])
            for inv in investigadores:
                writer.writerow([
                    inv.nombre, 
                    inv.email, 
                    inv.regional or "SANTANDER",
                    inv.sede or "CGAO VÉLEZ", 
                    inv.nivel_academico or "", 
                    inv.rol or "", 
                    "Activo" if inv.is_active else "Inactivo"
                ])
            output = BytesIO(str_output.getvalue().encode('utf-8-sig'))
            return StreamingResponse(
                output, 
                media_type="text/csv; charset=utf-8", 
                headers={"Content-Disposition": "attachment; filename=reporte_talento.csv"}
            )
    except (OperationalError, SQLAlchemyError) as db_err:
        raise db_err
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ─── 6. Estadísticas Resumen Dashboard ───────────────────────────────────────

@router.get("/estadisticas-resumen")
def get_estadisticas_resumen(
    current_user: User = Depends(get_current_investigador_or_instructor),
    db: Session = Depends(get_db)
):
    """Retorna estadísticas consolidadas para el dashboard de reportes."""
    try:
        total_proyectos = db.query(Proyecto).count()
        total_grupos = db.query(Grupo).count()
        total_semilleros = db.query(Semillero).count()
        total_productos = db.query(Producto).count()
        total_investigadores = db.query(User).filter(User.rol.in_(['investigador', 'instructor'])).count()
        
        proyectos_por_estado = {
            str(estado): int(count)
            for estado, count in db.query(Proyecto.estado, func.count(Proyecto.id))
            .filter(Proyecto.estado.isnot(None))
            .group_by(Proyecto.estado)
            .all()
        }
        
        proyectos_por_año = {
            str(año): int(count)
            for año, count in db.query(Proyecto.vigencia, func.count(Proyecto.id))
            .filter(Proyecto.vigencia.isnot(None))
            .group_by(Proyecto.vigencia)
            .all()
        }
        
        productos_por_tipo = {
            str(tipo): int(count)
            for tipo, count in db.query(Producto.tipo, func.count(Producto.id))
            .filter(Producto.tipo.isnot(None))
            .group_by(Producto.tipo)
            .all()
        }
        
        productos_verificados = db.query(Producto).filter(Producto.is_verificado == True).count()
        productos_pendientes = total_productos - productos_verificados
        
        grupos_por_clasificacion = {
            str(clasif): int(count)
            for clasif, count in db.query(Grupo.clasificacion, func.count(Grupo.id))
            .filter(Grupo.clasificacion.isnot(None))
            .group_by(Grupo.clasificacion)
            .all()
        }
        
        return {
            "totales": {
                "proyectos": total_proyectos,
                "grupos": total_grupos,
                "semilleros": total_semilleros,
                "productos": total_productos,
                "investigadores": total_investigadores
            },
            "proyectos_por_estado": proyectos_por_estado,
            "proyectos_por_año": proyectos_por_año,
            "productos_por_tipo": productos_por_tipo,
            "productos_verificacion": {
                "verificados": productos_verificados,
                "pendientes": productos_pendientes,
                "tasa_verificacion": round(productos_verificados / total_productos * 100, 2) if total_productos > 0 else 0
            },
            "grupos_por_clasificacion": grupos_por_clasificacion,
            "timestamp": datetime.now(timezone.utc).isoformat()
        }
    except (OperationalError, SQLAlchemyError) as db_err:
        raise db_err
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ─── 7. Certificado Individual de Investigador ───────────────────────────────

@router.get("/investigador/{user_id}/certificado")
def generar_certificado_investigador(
    user_id: str,
    current_user: User = Depends(get_current_investigador_or_instructor),
    db: Session = Depends(get_db)
):
    """Genera un certificado de participación en PDF para un investigador o instructor."""
    if current_user.rol != "admin" and str(current_user.id) != str(user_id):
        raise HTTPException(status_code=403, detail="No autorizado para descargar este certificado")
    try:
        user = db.query(User).filter(User.id == user_id).first()
        if not user:
            raise HTTPException(status_code=404, detail="Investigador no encontrado")

        try:
            from reportlab.pdfgen import canvas
            from reportlab.lib.pagesizes import letter
            from reportlab.lib.units import inch

            output = BytesIO()
            c = canvas.Canvas(output, pagesize=letter)
            width, height = letter

            c.setFont("Helvetica-Bold", 16)
            c.drawCentredString(width / 2.0, height - 100, "SENA - SERVICIO NACIONAL DE APRENDIZAJE")
            c.setFont("Helvetica-Bold", 14)
            c.drawCentredString(width / 2.0, height - 120, "CENTRO DE GESTIÓN AGROEMPRESARIAL Y ORIENTE - CGAO")
            c.drawCentredString(width / 2.0, height - 140, "SISTEMA SENNOVA")

            c.setFont("Helvetica", 12)
            c.drawCentredString(width / 2.0, height - 180, "CERTIFICA QUE:")

            c.setFont("Helvetica-Bold", 18)
            c.drawCentredString(width / 2.0, height - 220, user.nombre.upper())

            c.setFont("Helvetica", 12)
            c.drawCentredString(width / 2.0, height - 250, f"Identificado(a) con correo institucional {user.email}")

            textobject = c.beginText()
            textobject.setTextOrigin(inch, height - 300)
            textobject.setFont("Helvetica", 12)
            lines = [
                "Ha participado activamente como INVESTIGADOR en el ecosistema SENNOVA CGAO,",
                "liderando y apoyando proyectos de ciencia, tecnología e innovación.",
                "",
                "Válido para la vigencia actual."
            ]
            for line in lines:
                textobject.textLine(line)
            c.drawText(textobject)

            c.setFont("Helvetica", 10)
            c.drawString(inch, height - 400, f"Generado electrónicamente el: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            c.drawString(inch, height - 420, f"Código de Verificación: {str(uuid.uuid4())[:8]}")

            c.save()
            output.seek(0)
        except ImportError:
            output = BytesIO()
            content = f"Error: Se requiere la librería 'reportlab' para generar el PDF de {user.nombre}."
            output.write(content.encode('utf-8'))
            output.seek(0)

        return StreamingResponse(
            output,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=certificado_{user_id}.pdf"}
        )
    except (OperationalError, SQLAlchemyError) as db_err:
        raise db_err
    except Exception as e:
        if isinstance(e, HTTPException):
            raise e
        raise HTTPException(status_code=500, detail=str(e))
