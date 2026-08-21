"""
Tests dedicados para los generadores de Excel y CSV en SENNOVA
Valida la integridad binaria de los archivos .xlsx generados con openpyxl y los archivos .csv
"""

import os
import uuid
from datetime import date
from io import BytesIO
import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from db_support import db_path_for, sqlite_url_for

TEST_DB_FILE = db_path_for("test_sennova_excel_audit.db")
TEST_DB_URL = sqlite_url_for(TEST_DB_FILE)

os.environ["DATABASE_URL"] = TEST_DB_URL
os.environ["DEBUG"] = "true"

from app.database import Base, get_db
from app.main import app
from app.models import User, Grupo, Semillero, Proyecto, Producto
from app.auth import get_password_hash, create_access_token

test_engine = create_engine(TEST_DB_URL, connect_args={"check_same_thread": False})
TestingSessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=test_engine)

def override_get_db():
    db = TestingSessionLocal()
    try:
        yield db
    finally:
        db.close()


@pytest.fixture(scope="module", autouse=True)
def setup_excel_db():
    app.dependency_overrides[get_db] = override_get_db
    Base.metadata.drop_all(bind=test_engine)
    Base.metadata.create_all(bind=test_engine)
    
    db = TestingSessionLocal()
    inv = User(
        email="inv_excel@sena.edu.co",
        password_hash=get_password_hash("Inv123!"),
        nombre="Investigador Excel Test",
        rol="investigador",
        rol_sennova="Investigador Principal",
        sede="CGAO Vélez",
        regional="Santander",
        is_active=True
    )
    db.add(inv)
    db.commit()

    # Agregar grupo, semillero, proyecto, producto
    grupo = Grupo(
        nombre="Grupo Agroindustrial CGAO",
        codigo_gruplac="COL0012345",
        clasificacion="A1",
        owner_id=inv.id,
        lineas_investigacion=["Biotecnología", "Agroindustria"]
    )
    db.add(grupo)
    db.commit()

    semillero = Semillero(
        nombre="Semillero Biotecnología",
        sigla="SIBIOTIC",
        grupo_id=grupo.id,
        owner_id=inv.id,
        estado="Activo"
    )
    db.add(semillero)
    db.commit()

    proyecto = Proyecto(
        nombre="Plataforma de Monitoreo Agroecológico",
        nombre_corto="Monitoreo Agro",
        codigo_sgps="SGPS-2026-901",
        estado="En ejecución",
        vigencia=2026,
        presupuesto_total=35000000.0,
        tipologia="I+D",
        owner_id=inv.id
    )
    db.add(proyecto)
    db.commit()

    producto = Producto(
        tipo="Software",
        nombre="Software de Monitoreo IoT",
        descripcion="Sistema de monitoreo de sensores en tiempo real",
        fecha_publicacion=date(2026, 2, 15),
        doi="10.1000/182",
        is_verificado=True,
        proyecto_id=proyecto.id,
        owner_id=inv.id,
        url="https://github.com/sena/iot"
    )
    db.add(producto)
    db.commit()
    db.close()

    yield

    app.dependency_overrides.clear()
    Base.metadata.drop_all(bind=test_engine)
    test_engine.dispose()
    if os.path.exists(TEST_DB_FILE):
        try:
            os.remove(TEST_DB_FILE)
        except PermissionError:
            pass


@pytest.fixture(scope="module")
def client():
    return TestClient(app)


@pytest.fixture(scope="module")
def headers():
    db = TestingSessionLocal()
    inv = db.query(User).filter(User.email == "inv_excel@sena.edu.co").first()
    token = create_access_token(inv.id, inv.email, inv.rol)
    db.close()
    return {"Authorization": f"Bearer {token}"}


def test_excel_consolidado_proyectos(client, headers):
    """Prueba la generación de Excel consolidado de proyectos con openpyxl."""
    response = client.get("/reportes/proyectos-consolidado?formato=excel", headers=headers)
    assert response.status_code == 200
    assert "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in response.headers["content-type"]
    assert "consolidado_proyectos_" in response.headers.get("content-disposition", "")

    # Validar parseo binario con openpyxl
    wb = load_workbook(filename=BytesIO(response.content))
    assert "Consolidado Proyectos" in wb.sheetnames
    ws = wb["Consolidado Proyectos"]
    assert "CONSOLIDADO DE PROYECTOS" in ws["A1"].value
    assert ws.cell(row=4, column=1).value == "Código SGPS"
    assert ws.cell(row=4, column=6).value == "Presupuesto Total (COP)"
    assert ws.cell(row=5, column=1).value == "SGPS-2026-901"
    assert ws.cell(row=5, column=6).value == 35000000.0


def test_csv_consolidado_proyectos(client, headers):
    """Prueba la generación de CSV consolidado de proyectos con UTF-8-sig."""
    response = client.get("/reportes/proyectos-consolidado?formato=csv", headers=headers)
    assert response.status_code == 200
    assert "text/csv" in response.headers["content-type"]
    text = response.content.decode("utf-8-sig")
    assert "codigo_sgps" in text
    assert "SGPS-2026-901" in text
    assert "presupuesto_total_cop" in text


def test_excel_consolidado_grupos(client, headers):
    """Prueba la generación de Excel consolidado de grupos de investigación."""
    response = client.get("/reportes/grupos-consolidado?formato=excel", headers=headers)
    assert response.status_code == 200
    assert "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in response.headers["content-type"]

    wb = load_workbook(filename=BytesIO(response.content))
    assert "Grupos GRUPLAC" in wb.sheetnames
    ws = wb["Grupos GRUPLAC"]
    assert "GRUPOS DE INVESTIGACIÓN" in ws["A1"].value
    assert ws.cell(row=4, column=1).value == "Nombre del Grupo"
    assert ws.cell(row=5, column=1).value == "Grupo Agroindustrial CGAO"


def test_csv_consolidado_grupos(client, headers):
    """Prueba la generación de CSV consolidado de grupos."""
    response = client.get("/reportes/grupos-consolidado?formato=csv", headers=headers)
    assert response.status_code == 200
    text = response.content.decode("utf-8-sig")
    assert "nombre" in text
    assert "Grupo Agroindustrial CGAO" in text
    assert "COL0012345" in text


def test_excel_consolidado_productos(client, headers):
    """Prueba la generación de Excel consolidado de productos MinCiencias."""
    response = client.get("/reportes/productos-consolidado?formato=excel", headers=headers)
    assert response.status_code == 200
    assert "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in response.headers["content-type"]

    wb = load_workbook(filename=BytesIO(response.content))
    assert "Productos MinCiencias" in wb.sheetnames
    ws = wb["Productos MinCiencias"]
    assert "PRODUCTOS DE INVESTIGACIÓN" in ws["A1"].value
    assert ws.cell(row=4, column=1).value == "Tipo MinCiencias"
    assert ws.cell(row=5, column=1).value == "Software"


def test_csv_consolidado_productos(client, headers):
    """Prueba la generación de CSV consolidado de productos."""
    response = client.get("/reportes/productos-consolidado?formato=csv", headers=headers)
    assert response.status_code == 200
    text = response.content.decode("utf-8-sig")
    assert "tipo" in text
    assert "Software de Monitoreo IoT" in text
    assert "autor" in text


def test_excel_consolidado_semilleros(client, headers):
    """Prueba la generación de Excel consolidado de semilleros y aprendices."""
    response = client.get("/reportes/semilleros-consolidado?formato=excel", headers=headers)
    assert response.status_code == 200
    assert "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in response.headers["content-type"]

    wb = load_workbook(filename=BytesIO(response.content))
    assert "Semilleros y Aprendices" in wb.sheetnames
    ws = wb["Semilleros y Aprendices"]
    assert "SEMILLEROS DE INVESTIGACIÓN" in ws["A1"].value
    assert ws.cell(row=4, column=1).value == "Nombre del Semillero"
    assert ws.cell(row=5, column=1).value == "Semillero Biotecnología"


def test_csv_consolidado_semilleros(client, headers):
    """Prueba la generación de CSV consolidado de semilleros."""
    response = client.get("/reportes/semilleros-consolidado?formato=csv", headers=headers)
    assert response.status_code == 200
    text = response.content.decode("utf-8-sig")
    assert "nombre" in text
    assert "Semillero Biotecnología" in text


def test_excel_consolidado_talento(client, headers):
    """Prueba la generación de Excel consolidado de talento humano e investigadores."""
    response = client.get("/reportes/talento-consolidado?formato=excel", headers=headers)
    assert response.status_code == 200
    assert "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in response.headers["content-type"]

    wb = load_workbook(filename=BytesIO(response.content))
    assert "Talento SENNOVA" in wb.sheetnames
    ws = wb["Talento SENNOVA"]
    assert "DIRECTORIO CONSOLIDADO DE TALENTO HUMANO" in ws["A1"].value
    assert ws.cell(row=4, column=1).value == "Nombre Completo"
    assert ws.cell(row=5, column=1).value == "Investigador Excel Test"


def test_csv_consolidado_talento(client, headers):
    """Prueba la generación de CSV consolidado de talento humano."""
    response = client.get("/reportes/talento-consolidado?formato=csv", headers=headers)
    assert response.status_code == 200
    text = response.content.decode("utf-8-sig")
    assert "nombre" in text
    assert "Investigador Excel Test" in text
    assert "regional" in text
